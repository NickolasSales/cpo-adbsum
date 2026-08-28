"""
Importacao de alunos a partir de CSV ou XLSX.

O fluxo tem duas fases estritamente separadas:

    analisar()   le o arquivo, confronta com o banco e classifica cada linha.
                 NAO escreve nada. Nenhum User, StudentProfile ou Enrollment
                 e criado, nenhuma senha e definida.

    confirmar()  reexecuta a analise a partir das linhas cruas e aplica
                 somente o que for importavel, dentro de uma unica transacao.

A reexecucao na confirmacao e proposital. O preview renderizado e informativo,
nunca autoritativo: entre o preview e a confirmacao o banco pode ter mudado —
um modulo pode ter sido desativado, um aluno pode ter sido criado por outra
via. Nenhum identificador vindo do navegador participa da decisao.
"""

import csv
import io
import logging
import unicodedata
import zipfile
from dataclasses import dataclass, field
from xml.etree.ElementTree import ParseError

from django.db import transaction

from accounts.managers import normalizar_email
from accounts.models import User, UserRole
from audit.models import AuditEvent
from audit.services import record
from common.exceptions import DomainError
from courses.models import Enrollment, EnrollmentStatus, Module, normalizar_codigo
from courses.services import create_enrollment
from students.models import StudentSource
from students.services import create_student, obter_senha_inicial

logger = logging.getLogger("cpo.students")

# Esta aplicacao lida com centenas de alunos, nao com milhoes. Limites
# apertados sao uma protecao barata contra upload abusivo e contra sessoes
# gigantes.
TAMANHO_MAXIMO_BYTES = 5 * 1024 * 1024
MAX_LINHAS = 2000

EXTENSOES_ACEITAS = (".csv", ".xlsx")

COLUNAS_OBRIGATORIAS = ("nome", "email", "modulo")

# Cabecalhos alternativos aceitos, ja normalizados (minusculas, sem acento).
ALIAS_COLUNAS = {
    "nome": "nome",
    "nome completo": "nome",
    "aluno": "nome",
    "email": "email",
    "e-mail": "email",
    "modulo": "modulo",
    "modulos": "modulo",
    "codigo": "modulo",
    "codigo do modulo": "modulo",
}


class RowStatus:
    """Classificacao de cada linha do arquivo."""

    NOVO_ALUNO = "NOVO ALUNO"
    ALUNO_EXISTENTE = "ALUNO EXISTENTE"
    MATRICULA_EXISTENTE = "MATRICULA EXISTENTE"
    MATRICULA_INATIVA = "MATRICULA INATIVA EXISTENTE"
    LINHA_DUPLICADA = "LINHA DUPLICADA"
    MODULO_NAO_ENCONTRADO = "MODULO NAO ENCONTRADO"
    EMAIL_INVALIDO = "E-MAIL INVALIDO"
    NOME_AUSENTE = "NOME AUSENTE"
    MODULO_AUSENTE = "MODULO AUSENTE"
    EMAIL_DE_ADMIN = "E-MAIL DE ADMINISTRADOR"


# Linhas que produzem escrita no banco quando o administrador confirma.
STATUS_IMPORTAVEIS = frozenset({RowStatus.NOVO_ALUNO, RowStatus.ALUNO_EXISTENTE})

# Linhas que impedem a importacao daquela linha por erro de conteudo.
STATUS_INVALIDOS = frozenset(
    {
        RowStatus.EMAIL_INVALIDO,
        RowStatus.NOME_AUSENTE,
        RowStatus.MODULO_AUSENTE,
        RowStatus.MODULO_NAO_ENCONTRADO,
        RowStatus.EMAIL_DE_ADMIN,
    }
)

# Linhas ignoradas sem que isso seja erro: o estado desejado ja existe.
STATUS_IGNORADOS = frozenset(
    {
        RowStatus.MATRICULA_EXISTENTE,
        RowStatus.MATRICULA_INATIVA,
        RowStatus.LINHA_DUPLICADA,
    }
)


# ---------------------------------------------------------------------------
# Leitura do arquivo
# ---------------------------------------------------------------------------


def sem_acento(texto):
    normalizado = unicodedata.normalize("NFKD", texto or "")
    return "".join(c for c in normalizado if not unicodedata.combining(c))


def normalizar_cabecalho(valor):
    return sem_acento(str(valor or "")).strip().lower()


def validar_upload(arquivo):
    """
    Valida o arquivo antes de qualquer tentativa de leitura.

    O nome do arquivo nao e prova de nada: a extensao e apenas o primeiro
    filtro, e a estrutura real e conferida na leitura. O tamanho e checado
    antes de tudo para nao carregar um arquivo grande em memoria.
    """
    if arquivo is None:
        raise DomainError("Selecione um arquivo para importar.")

    nome = (getattr(arquivo, "name", "") or "").lower()
    if not nome.endswith(EXTENSOES_ACEITAS):
        raise DomainError(
            "Formato nao suportado. Envie um arquivo .csv ou .xlsx."
        )

    tamanho = getattr(arquivo, "size", 0) or 0
    if tamanho > TAMANHO_MAXIMO_BYTES:
        raise DomainError(
            "O arquivo excede o limite de {} MB.".format(
                TAMANHO_MAXIMO_BYTES // (1024 * 1024)
            )
        )
    if tamanho == 0:
        raise DomainError("O arquivo esta vazio.")

    return nome


def ler_arquivo(arquivo):
    """Devolve as linhas cruas do arquivo como dicionarios nome/email/modulo."""
    nome = validar_upload(arquivo)
    arquivo.seek(0)

    if nome.endswith(".csv"):
        return _ler_csv(arquivo)
    return _ler_xlsx(arquivo)


def _decodificar_csv(bruto):
    """
    Decodifica o conteudo do CSV.

    utf-8-sig cobre UTF-8 puro e UTF-8 com BOM, que e o que o Excel grava ao
    escolher "CSV UTF-8". Nao ha fallback silencioso para cp1252: aceitar
    qualquer byte produziria acentuacao corrompida sem aviso, e um nome
    gravado errado no certificado e pior do que uma recusa clara.
    """
    try:
        return bruto.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise DomainError(
            "Nao foi possivel ler o arquivo: a codificacao nao e UTF-8. "
            "Salve novamente como 'CSV UTF-8' ou envie um arquivo .xlsx."
        )


def _detectar_delimitador(amostra):
    """
    Descobre o separador entre virgula e ponto e virgula.

    O Excel em portugues grava CSV com ponto e virgula. Aceitar os dois evita
    que o arquivo mais comum do usuario seja lido como uma unica coluna.
    """
    try:
        return csv.Sniffer().sniff(amostra, delimiters=",;").delimiter
    except csv.Error:
        return ";" if amostra.count(";") > amostra.count(",") else ","


def _ler_csv(arquivo):
    texto = _decodificar_csv(arquivo.read())
    if not texto.strip():
        raise DomainError("O arquivo esta vazio.")

    delimitador = _detectar_delimitador(texto[:4096])
    leitor = csv.reader(io.StringIO(texto), delimiter=delimitador)

    try:
        linhas = list(leitor)
    except csv.Error as exc:
        raise DomainError("Nao foi possivel interpretar o CSV: {}".format(exc))

    return _montar_linhas(linhas)


# Excecoes que significam, todas elas, a mesma coisa: este arquivo nao serve.
#
# A lista e explicita de proposito. Uma captura ampla engoliria tambem
# TypeError ou AttributeError vindos de um defeito nosso, transformando um bug
# de programacao numa mensagem amigavel — exatamente o que nao pode acontecer
# durante o desenvolvimento.
#
# Cada entrada foi verificada contra o openpyxl 3.1 com um arquivo real:
#
#   BadZipFile   texto puro, arquivo vazio, .xlsx truncado, .xls antigo ou
#                planilha protegida por senha (que e um container OLE2, nao
#                um zip). Nao herda de OSError — este foi o furo original.
#   KeyError     zip integro sem as pecas internas de um xlsx
#   OSError      pacote sem os relacionamentos que apontam para a planilha
#   ParseError   XML malformado dentro do zip
#   IndexError   pacote valido que nao contem nenhuma aba
#   ValueError   valores fora de faixa nas estruturas internas
#   InvalidFileException  formato que o openpyxl recusa de saida (adicionada
#                em _ler_xlsx, junto do import sob demanda)
#
# Observacao deliberada sobre dependencias: com o lxml instalado o openpyxl
# troca de parser e o XML malformado passa a levantar lxml.etree.XMLSyntaxError,
# que nao herda de ParseError. O lxml nao esta em requirements.txt, e esta nota
# existe para que ninguem o acrescente sem rever esta tupla.
ERROS_DE_PLANILHA_INVALIDA = (
    zipfile.BadZipFile,
    KeyError,
    OSError,
    ValueError,
    ParseError,
    IndexError,
)


def _ler_xlsx(arquivo):
    # Importado sob demanda para que o openpyxl so seja carregado quando um
    # XLSX for de fato enviado.
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    invalidos = ERROS_DE_PLANILHA_INVALIDA + (InvalidFileException,)
    planilha = None

    try:
        planilha = load_workbook(
            filename=io.BytesIO(arquivo.read()), read_only=True, data_only=True
        )
        # A leitura das linhas fica dentro do mesmo try porque, em read_only,
        # o openpyxl adia o parsing: um sheet1.xml corrompido atravessa o
        # load_workbook inteiro e so estoura aqui, na iteracao.
        aba = planilha.worksheets[0]
        linhas = [
            ["" if celula is None else celula for celula in linha]
            for linha in aba.iter_rows(values_only=True)
        ]
    except invalidos as exc:
        logger.info("Planilha recusada na leitura (%s)", exc.__class__.__name__)
        raise DomainError(
            "Nao foi possivel ler a planilha. Confirme que o arquivo e um "
            ".xlsx valido e nao esta protegido por senha."
        )
    finally:
        if planilha is not None:
            planilha.close()

    return _montar_linhas(linhas)


def _texto(valor):
    """
    Converte uma celula em texto limpo.

    Numeros inteiros vindos do Excel chegam como float (1.0). Sem este
    tratamento um codigo de modulo numerico viraria "1.0".
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).strip()


def _montar_linhas(linhas_brutas):
    """Mapeia as colunas do cabecalho e devolve as linhas de dados."""
    if not linhas_brutas:
        raise DomainError("O arquivo esta vazio.")

    cabecalho = linhas_brutas[0]
    indices = {}
    for posicao, titulo in enumerate(cabecalho):
        chave = ALIAS_COLUNAS.get(normalizar_cabecalho(titulo))
        # A primeira ocorrencia vence, para que uma coluna repetida no
        # arquivo nao mude qual delas e lida.
        if chave and chave not in indices:
            indices[chave] = posicao

    faltando = [c for c in COLUNAS_OBRIGATORIAS if c not in indices]
    if faltando:
        raise DomainError(
            "O arquivo precisa conter as colunas: {}. Faltando: {}.".format(
                ", ".join(COLUNAS_OBRIGATORIAS), ", ".join(faltando)
            )
        )

    linhas = []
    for numero, bruta in enumerate(linhas_brutas[1:], start=2):
        def coluna(chave):
            posicao = indices[chave]
            return _texto(bruta[posicao]) if posicao < len(bruta) else ""

        nome, email, modulo = coluna("nome"), coluna("email"), coluna("modulo")

        # Linha totalmente em branco e ruido de planilha, nao erro do usuario.
        if not any((nome, email, modulo)):
            continue

        linhas.append(
            {"numero": numero, "nome": nome, "email": email, "modulo": modulo}
        )

    if not linhas:
        raise DomainError("O arquivo nao possui nenhuma linha de dados.")

    if len(linhas) > MAX_LINHAS:
        raise DomainError(
            "O arquivo possui {} linhas e o limite por importacao e {}. "
            "Divida o arquivo em partes menores.".format(len(linhas), MAX_LINHAS)
        )

    return linhas


# ---------------------------------------------------------------------------
# Analise
# ---------------------------------------------------------------------------


@dataclass
class LinhaAnalisada:
    numero: int
    nome: str
    email: str
    modulo: str
    status: str
    aviso: str = ""

    @property
    def importavel(self):
        return self.status in STATUS_IMPORTAVEIS

    @property
    def invalida(self):
        return self.status in STATUS_INVALIDOS


@dataclass
class ResultadoAnalise:
    linhas: list = field(default_factory=list)

    @property
    def total_linhas(self):
        return len(self.linhas)

    @property
    def importaveis(self):
        return [linha for linha in self.linhas if linha.importavel]

    @property
    def novos_alunos(self):
        """Alunos distintos a criar; o mesmo e-mail em dois modulos conta uma vez."""
        return len(
            {
                linha.email
                for linha in self.linhas
                if linha.status == RowStatus.NOVO_ALUNO
            }
        )

    @property
    def alunos_existentes(self):
        return len(
            {
                linha.email
                for linha in self.linhas
                if linha.status == RowStatus.ALUNO_EXISTENTE
            }
        )

    @property
    def novas_matriculas(self):
        return len(self.importaveis)

    @property
    def linhas_invalidas(self):
        return len([linha for linha in self.linhas if linha.invalida])

    @property
    def modulo_inexistente(self):
        return len(
            [
                linha
                for linha in self.linhas
                if linha.status == RowStatus.MODULO_NAO_ENCONTRADO
            ]
        )

    @property
    def matriculas_existentes(self):
        return len(
            [
                linha
                for linha in self.linhas
                if linha.status
                in (RowStatus.MATRICULA_EXISTENTE, RowStatus.MATRICULA_INATIVA)
            ]
        )

    @property
    def linhas_duplicadas(self):
        return len(
            [linha for linha in self.linhas if linha.status == RowStatus.LINHA_DUPLICADA]
        )

    @property
    def ignoradas(self):
        return len([linha for linha in self.linhas if linha.status in STATUS_IGNORADOS])

    @property
    def tem_algo_a_importar(self):
        return bool(self.importaveis)

    def resumo(self):
        return {
            "total_linhas": self.total_linhas,
            "novos_alunos": self.novos_alunos,
            "alunos_existentes": self.alunos_existentes,
            "novas_matriculas": self.novas_matriculas,
            "linhas_invalidas": self.linhas_invalidas,
            "modulo_inexistente": self.modulo_inexistente,
            "matriculas_existentes": self.matriculas_existentes,
            "linhas_duplicadas": self.linhas_duplicadas,
        }


def _email_valido(email):
    """
    Validacao deliberadamente conservadora.

    Nao tenta implementar a RFC: exige exatamente um arroba, com conteudo dos
    dois lados e um ponto no dominio. Rejeitar um endereco exotico e valido e
    preferivel a aceitar um endereco digitado errado, que geraria uma conta
    inacessivel.
    """
    if not email or email.count("@") != 1:
        return False
    local, _, dominio = email.partition("@")
    if not local or not dominio or " " in email:
        return False
    return "." in dominio and not dominio.startswith(".") and not dominio.endswith(".")


def _indexar_modulos():
    """
    Indice de modulos por codigo e por nome.

    O nome so entra no indice quando for inequivoco: se dois modulos tiverem
    o mesmo nome, nenhum dos dois e resolvivel por nome, e a linha cai em
    MODULO NAO ENCONTRADO. Escolher um deles arbitrariamente matricularia o
    aluno no modulo errado em silencio.
    """
    modulos = list(Module.objects.all())

    por_codigo = {modulo.code: modulo for modulo in modulos}

    contagem_por_nome = {}
    for modulo in modulos:
        chave = sem_acento(modulo.name).strip().lower()
        contagem_por_nome[chave] = contagem_por_nome.get(chave, 0) + 1

    por_nome = {}
    for modulo in modulos:
        chave = sem_acento(modulo.name).strip().lower()
        if contagem_por_nome[chave] == 1:
            por_nome[chave] = modulo

    return por_codigo, por_nome


def analisar(linhas_cruas):
    """
    Classifica cada linha confrontando o arquivo com o banco.

    Somente leitura. Nenhuma escrita acontece aqui, em nenhuma circunstancia.
    """
    por_codigo, por_nome = _indexar_modulos()

    emails = {
        normalizar_email(linha.get("email", "")) for linha in linhas_cruas
    }
    emails.discard("")
    usuarios = {
        user.email: user for user in User.objects.filter(email__in=emails)
    }

    # Uma unica consulta traz todas as matriculas relevantes, evitando N+1.
    matriculas = {}
    ids_usuarios = [user.pk for user in usuarios.values()]
    if ids_usuarios:
        for matricula in Enrollment.objects.filter(
            student_id__in=ids_usuarios
        ).select_related("module"):
            matriculas[(matricula.student_id, matricula.module_id)] = matricula

    vistos = set()
    analisadas = []

    for linha in linhas_cruas:
        numero = linha.get("numero", 0)
        nome = (linha.get("nome") or "").strip()
        email = normalizar_email(linha.get("email", ""))
        codigo = (linha.get("modulo") or "").strip()

        def montar(status, aviso=""):
            return LinhaAnalisada(
                numero=numero,
                nome=nome,
                email=email,
                modulo=codigo,
                status=status,
                aviso=aviso,
            )

        if not _email_valido(email):
            analisadas.append(montar(RowStatus.EMAIL_INVALIDO))
            continue
        if not nome:
            analisadas.append(montar(RowStatus.NOME_AUSENTE))
            continue
        if not codigo:
            analisadas.append(montar(RowStatus.MODULO_AUSENTE))
            continue

        modulo = por_codigo.get(normalizar_codigo(codigo))
        if modulo is None:
            modulo = por_nome.get(sem_acento(codigo).strip().lower())
        if modulo is None:
            analisadas.append(montar(RowStatus.MODULO_NAO_ENCONTRADO))
            continue

        # Duplicidade dentro do proprio arquivo: a primeira ocorrencia segue,
        # as demais sao marcadas e nao geram escrita.
        chave = (email, modulo.pk)
        if chave in vistos:
            analisadas.append(montar(RowStatus.LINHA_DUPLICADA))
            continue
        vistos.add(chave)

        usuario = usuarios.get(email)

        if usuario is None:
            if not modulo.is_active:
                analisadas.append(
                    montar(
                        RowStatus.MODULO_NAO_ENCONTRADO,
                        aviso="O modulo {} esta inativo.".format(modulo.code),
                    )
                )
                continue
            analisadas.append(montar(RowStatus.NOVO_ALUNO))
            continue

        if usuario.role == UserRole.ADMIN:
            # Converter um administrador em aluno automaticamente seria uma
            # mudanca de privilegio silenciosa. A linha e recusada.
            analisadas.append(montar(RowStatus.EMAIL_DE_ADMIN))
            continue

        matricula = matriculas.get((usuario.pk, modulo.pk))
        if matricula is not None:
            if matricula.status == EnrollmentStatus.ACTIVE:
                analisadas.append(montar(RowStatus.MATRICULA_EXISTENTE))
            else:
                # Reativacao nunca acontece por importacao: e uma decisao
                # administrativa consciente, feita na tela de matriculas.
                analisadas.append(
                    montar(
                        RowStatus.MATRICULA_INATIVA,
                        aviso="Reative pela tela de matriculas, se desejado.",
                    )
                )
            continue

        if not modulo.is_active:
            analisadas.append(
                montar(
                    RowStatus.MODULO_NAO_ENCONTRADO,
                    aviso="O modulo {} esta inativo.".format(modulo.code),
                )
            )
            continue

        # A importacao matricula; nao sobrescreve cadastro existente. Se o
        # nome no arquivo diverge, o administrador e avisado e decide o que
        # fazer pela tela de edicao.
        aviso = ""
        if nome and nome.strip().lower() != usuario.full_name.strip().lower():
            aviso = "Nome no arquivo difere do cadastro atual ({}).".format(
                usuario.full_name
            )
        analisadas.append(montar(RowStatus.ALUNO_EXISTENTE, aviso=aviso))

    return ResultadoAnalise(linhas=analisadas)


# ---------------------------------------------------------------------------
# Confirmacao
# ---------------------------------------------------------------------------


@transaction.atomic
def confirmar(linhas_cruas, *, actor=None, request=None):
    """
    Aplica a importacao.

    A analise e refeita aqui a partir das linhas cruas: o que foi renderizado
    no preview nao decide nada. Somente linhas classificadas como importaveis
    neste instante produzem escrita.

    Todo o lote roda em uma unica transacao. Se qualquer linha falhar de forma
    inesperada, nada e gravado — nao existe importacao pela metade.
    """
    # Falha cedo, antes de qualquer escrita, se a senha inicial nao estiver
    # configurada.
    obter_senha_inicial()

    analise = analisar(linhas_cruas)
    if not analise.tem_algo_a_importar:
        raise DomainError("Nao ha nenhuma linha valida para importar.")

    por_codigo, por_nome = _indexar_modulos()
    criados_por_email = {}
    alunos_criados = 0
    matriculas_criadas = 0

    for linha in analise.importaveis:
        modulo = por_codigo.get(normalizar_codigo(linha.modulo))
        if modulo is None:
            modulo = por_nome.get(sem_acento(linha.modulo).strip().lower())

        usuario = criados_por_email.get(linha.email)
        if usuario is None:
            usuario = User.objects.filter(email=linha.email).first()

        if usuario is None:
            usuario = create_student(
                full_name=linha.nome,
                email=linha.email,
                source=StudentSource.IMPORT,
                actor=actor,
                request=request,
            )
            criados_por_email[linha.email] = usuario
            alunos_criados += 1

        create_enrollment(
            student=usuario, module=modulo, actor=actor, request=request
        )
        matriculas_criadas += 1

    record(
        AuditEvent.STUDENT_IMPORT_COMPLETED,
        request=request,
        actor=actor,
        metadata={
            # Apenas contagens. A lista de e-mails nao vai para a trilha:
            # os eventos individuais ja dao a rastreabilidade por aluno.
            "rows_received": analise.total_linhas,
            "students_created": alunos_criados,
            "enrollments_created": matriculas_criadas,
            "skipped": analise.ignoradas,
            "invalid": analise.linhas_invalidas,
        },
    )

    return {
        "alunos_criados": alunos_criados,
        "matriculas_criadas": matriculas_criadas,
        "ignoradas": analise.ignoradas,
        "invalidas": analise.linhas_invalidas,
        "total_linhas": analise.total_linhas,
    }
