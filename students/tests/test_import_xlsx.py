"""Importacao via XLSX e o fluxo completo de upload, preview e confirmacao."""

import io
from unittest import mock

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

from accounts.models import User, UserRole
from audit.models import AuditEvent, AuditLog
from common.exceptions import DomainError
from courses.models import Enrollment
from students.importers import RowStatus, analisar, ler_arquivo
from students.models import StudentProfile, StudentSource

pytestmark = pytest.mark.django_db

URL_UPLOAD = "/admin-panel/alunos/importar/"
URL_PREVIEW = "/admin-panel/alunos/importar/preview/"
URL_CONFIRMAR = "/admin-panel/alunos/importar/confirmar/"
URL_CANCELAR = "/admin-panel/alunos/importar/cancelar/"


def planilha(linhas, nome="alunos.xlsx"):
    """Monta um .xlsx em memoria a partir de uma lista de listas."""
    from openpyxl import Workbook

    livro = Workbook()
    aba = livro.active
    for linha in linhas:
        aba.append(linha)

    buffer = io.BytesIO()
    livro.save(buffer)
    return SimpleUploadedFile(
        nome,
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def csv(texto, nome="alunos.csv"):
    return SimpleUploadedFile(nome, texto.encode("utf-8"), content_type="text/csv")


def contagens():
    return (
        User.objects.count(),
        StudentProfile.objects.count(),
        Enrollment.objects.count(),
    )


# ---------------------------------------------------------------------------
# Leitura do XLSX
# ---------------------------------------------------------------------------


def test_xlsx_valido_e_lido():
    arquivo = planilha(
        [
            ["nome", "email", "modulo"],
            ["Joao da Silva", "joao@exemplo.test", "MOD1"],
            ["Maria Oliveira", "maria@exemplo.test", "MOD2"],
        ]
    )
    linhas = ler_arquivo(arquivo)

    assert len(linhas) == 2
    assert linhas[0]["nome"] == "Joao da Silva"
    assert linhas[0]["email"] == "joao@exemplo.test"
    assert linhas[0]["modulo"] == "MOD1"
    assert linhas[0]["numero"] == 2


def test_xlsx_aceita_cabecalhos_em_caixa_variada():
    arquivo = planilha(
        [["Nome", "E-MAIL", "Modulo"], ["Ana", "ana@exemplo.test", "MOD1"]]
    )
    linhas = ler_arquivo(arquivo)
    assert linhas[0]["nome"] == "Ana"
    assert linhas[0]["email"] == "ana@exemplo.test"


def test_xlsx_remove_espacos_em_volta_dos_valores():
    arquivo = planilha(
        [["nome", "email", "modulo"], ["  Joao da Silva  ", "  joao@exemplo.test ", " MOD1 "]]
    )
    linhas = ler_arquivo(arquivo)
    assert linhas[0]["nome"] == "Joao da Silva"
    assert linhas[0]["modulo"] == "MOD1"


def test_xlsx_converte_numero_inteiro_sem_casa_decimal():
    """
    O Excel entrega numeros como float. Sem tratamento, um codigo de modulo
    numerico viraria "1.0" e nunca casaria com o cadastro.
    """
    arquivo = planilha([["nome", "email", "modulo"], ["Ana", "ana@exemplo.test", 1]])
    linhas = ler_arquivo(arquivo)
    assert linhas[0]["modulo"] == "1"


def test_xlsx_ignora_linhas_totalmente_em_branco():
    arquivo = planilha(
        [
            ["nome", "email", "modulo"],
            ["Ana", "ana@exemplo.test", "MOD1"],
            [None, None, None],
            ["Bruno", "bruno@exemplo.test", "MOD1"],
        ]
    )
    assert len(ler_arquivo(arquivo)) == 2


def test_xlsx_sem_coluna_obrigatoria():
    arquivo = planilha([["nome", "email"], ["Ana", "ana@exemplo.test"]])
    with pytest.raises(DomainError) as erro:
        ler_arquivo(arquivo)
    assert "modulo" in str(erro.value)


def test_xlsx_somente_com_cabecalho():
    with pytest.raises(DomainError):
        ler_arquivo(planilha([["nome", "email", "modulo"]]))


def test_arquivo_que_nao_e_xlsx_de_verdade():
    falso = SimpleUploadedFile(
        "alunos.xlsx", b"isto nao e uma planilha", content_type="application/vnd.ms-excel"
    )
    with pytest.raises(DomainError) as erro:
        ler_arquivo(falso)
    assert "planilha" in str(erro.value).lower()


# ---------------------------------------------------------------------------
# Planilhas corrompidas
#
# Cada teste abaixo corresponde a uma excecao concreta da tupla
# ERROS_DE_PLANILHA_INVALIDA. Todos precisam terminar em DomainError: o
# administrador recebe uma mensagem, nunca um erro 500.
# ---------------------------------------------------------------------------


def xlsx_adulterado(mutador, nome="alunos.xlsx"):
    """
    Reescreve um .xlsx valido peca por peca, aplicando o mutador a cada uma.

    O mutador recebe (nome_interno, conteudo) e devolve o mesmo par; devolver
    nome None remove aquela peca do pacote.
    """
    import zipfile

    original = planilha([["nome", "email", "modulo"], ["Ana", "ana@exemplo.test", "MOD1"]])
    entrada = zipfile.ZipFile(io.BytesIO(original.read()))

    saida = io.BytesIO()
    with zipfile.ZipFile(saida, "w") as pacote:
        for interno in entrada.namelist():
            novo_nome, dados = mutador(interno, entrada.read(interno))
            if novo_nome is not None:
                pacote.writestr(novo_nome, dados)

    return SimpleUploadedFile(nome, saida.getvalue(), content_type="application/vnd.ms-excel")


def test_arquivo_vazio_e_recusado():
    vazio = SimpleUploadedFile("alunos.xlsx", b"", content_type="application/vnd.ms-excel")
    with pytest.raises(DomainError):
        ler_arquivo(vazio)


def test_planilha_protegida_por_senha_e_recusada():
    """
    Um .xlsx protegido por senha e um container OLE2, nao um zip.

    E o caso que originou a correcao: zipfile.BadZipFile nao herda de OSError,
    entao escapava da captura antiga e virava erro 500.
    """
    ole2 = SimpleUploadedFile(
        "alunos.xlsx",
        bytes([0xD0, 0xCF, 0x11, 0xE0, 0xA1, 0xB1, 0x1A, 0xE1]) + bytes(512),
        content_type="application/vnd.ms-excel",
    )
    with pytest.raises(DomainError) as erro:
        ler_arquivo(ole2)
    assert "senha" in str(erro.value).lower()


def test_zip_sem_as_pecas_de_um_xlsx_e_recusado():
    """Zip integro, conteudo errado: o openpyxl levanta KeyError."""
    with pytest.raises(DomainError):
        ler_arquivo(xlsx_adulterado(lambda nome, dados: (None, dados)))


def test_pacote_sem_a_planilha_e_recusado():
    """Sem sheet1.xml o openpyxl abre o arquivo e depois nao encontra aba."""
    with pytest.raises(DomainError):
        ler_arquivo(
            xlsx_adulterado(
                lambda nome, dados: (None, dados) if nome.endswith("sheet1.xml") else (nome, dados)
            )
        )


def test_planilha_com_xml_truncado_e_recusada():
    """
    Corrupcao que so aparece durante a iteracao.

    Em read_only o openpyxl adia o parsing, entao este arquivo atravessa o
    load_workbook inteiro e so quebra na leitura das linhas. Sem o try
    envolvendo tambem a iteracao, o resultado seria um erro 500.
    """
    def truncar(nome, dados):
        return (nome, dados[: len(dados) // 2] if nome.endswith("sheet1.xml") else dados)

    with pytest.raises(DomainError):
        ler_arquivo(xlsx_adulterado(truncar))


def test_erro_de_programacao_nao_vira_mensagem_amigavel():
    """
    Contraprova da captura explicita.

    Um TypeError vindo de um defeito nosso precisa continuar estourando e
    aparecendo nos testes. Se ele virasse DomainError, o bug ficaria
    escondido atras de "confirme que o arquivo e um .xlsx valido".
    """
    with mock.patch("openpyxl.load_workbook", side_effect=TypeError("defeito interno")):
        with pytest.raises(TypeError):
            ler_arquivo(planilha([["nome", "email", "modulo"], ["Ana", "a@e.test", "MOD1"]]))


def test_upload_de_planilha_corrompida_responde_com_mensagem(admin_client_logado, modulo):
    """A recusa chega ao administrador como tela, nao como erro do servidor."""
    antes = contagens()
    resposta = admin_client_logado.post(
        URL_UPLOAD,
        {
            "arquivo": SimpleUploadedFile(
                "alunos.xlsx", b"isto nao e uma planilha", content_type="application/vnd.ms-excel"
            )
        },
    )

    assert resposta.status_code == 200
    assert contagens() == antes


# ---------------------------------------------------------------------------
# Analise a partir de XLSX
# ---------------------------------------------------------------------------


def test_analise_de_xlsx_classifica_os_mesmos_cenarios(
    modulo, outro_modulo, student_user, admin_user
):
    arquivo = planilha(
        [
            ["nome", "email", "modulo"],
            ["Novo Aluno", "novo@exemplo.test", "MOD1"],
            [student_user.full_name, student_user.email, "MOD2"],
            ["Sem Modulo", "semmod@exemplo.test", "MOD404"],
            ["Email Ruim", "nao-e-email", "MOD1"],
            ["Novo Aluno", "novo@exemplo.test", "MOD1"],
            ["Um Admin", admin_user.email, "MOD1"],
        ]
    )
    analise = analisar(ler_arquivo(arquivo))
    status = [linha.status for linha in analise.linhas]

    assert status == [
        RowStatus.NOVO_ALUNO,
        RowStatus.ALUNO_EXISTENTE,
        RowStatus.MODULO_NAO_ENCONTRADO,
        RowStatus.EMAIL_INVALIDO,
        RowStatus.LINHA_DUPLICADA,
        RowStatus.EMAIL_DE_ADMIN,
    ]


def test_analise_nao_escreve_no_banco(modulo):
    antes = contagens()
    arquivo = planilha(
        [["nome", "email", "modulo"], ["Novo Aluno", "novo@exemplo.test", "MOD1"]]
    )
    analisar(ler_arquivo(arquivo))
    assert contagens() == antes


# ---------------------------------------------------------------------------
# Fluxo pelas views
# ---------------------------------------------------------------------------


CSV_EXEMPLO = (
    "nome,email,modulo\n"
    "Joao da Silva,joao@example.com,MOD1\n"
    "Maria Oliveira,maria@example.com,MOD1\n"
    "Pedro Souza,pedro@example.com,MOD2\n"
    "Joao da Silva,joao@example.com,MOD2\n"
    "Joao da Silva,joao@example.com,MOD1\n"
    "Sem Modulo,semmod@example.com,MOD404\n"
)


def enviar(client, arquivo):
    return client.post(URL_UPLOAD, {"arquivo": arquivo})


def test_upload_nao_altera_o_banco(admin_client_logado, modulo, outro_modulo):
    """
    Requisito central da importacao: o preview e somente leitura.

    Nenhum User, StudentProfile ou Enrollment pode existir antes de o
    administrador confirmar.
    """
    antes = contagens()
    resposta = enviar(admin_client_logado, csv(CSV_EXEMPLO))

    assert resposta.status_code == 302
    assert resposta.url == URL_PREVIEW
    assert contagens() == antes


def test_preview_traz_o_resumo(admin_client_logado, modulo, outro_modulo):
    enviar(admin_client_logado, csv(CSV_EXEMPLO))
    resposta = admin_client_logado.get(URL_PREVIEW)

    assert resposta.status_code == 200
    resumo = resposta.context["resumo"]
    assert resumo["total_linhas"] == 6
    assert resumo["novos_alunos"] == 3
    assert resumo["novas_matriculas"] == 4
    assert resumo["linhas_duplicadas"] == 1
    assert resumo["modulo_inexistente"] == 1
    assert resposta.context["token"]
    assert len(resposta.context["linhas"]) == 6


def test_preview_sem_upload_previo_volta_para_o_envio(admin_client_logado):
    resposta = admin_client_logado.get(URL_PREVIEW)
    assert resposta.status_code == 302
    assert resposta.url == URL_UPLOAD


def test_confirmacao_importa(admin_client_logado, modulo, outro_modulo, senha_padrao):
    enviar(admin_client_logado, csv(CSV_EXEMPLO))
    token = admin_client_logado.get(URL_PREVIEW).context["token"]

    resposta = admin_client_logado.post(URL_CONFIRMAR, {"token": token})
    assert resposta.status_code == 302

    alunos = User.objects.filter(role=UserRole.STUDENT)
    assert alunos.count() == 3
    assert StudentProfile.objects.count() == 3
    assert Enrollment.objects.count() == 4

    joao = User.objects.get(email="joao@example.com")
    # Etapa 5: sem troca obrigatoria. A senha da importacao continua sendo a
    # padrao do ambiente, e o administrador pode redefini-la depois.
    assert joao.must_change_password is False
    assert joao.is_active is True
    assert joao.student_profile.source == StudentSource.IMPORT
    assert joao.password.startswith("pbkdf2_")
    assert joao.check_password(senha_padrao) is True
    # Mesmo aluno em dois modulos: uma conta, duas matriculas.
    assert joao.enrollments.count() == 2


def test_confirmacao_com_token_errado_nao_importa(
    admin_client_logado, modulo, outro_modulo
):
    enviar(admin_client_logado, csv(CSV_EXEMPLO))
    admin_client_logado.get(URL_PREVIEW)
    antes = contagens()

    resposta = admin_client_logado.post(URL_CONFIRMAR, {"token": "token-invalido"})
    assert resposta.status_code == 302
    assert contagens() == antes


def test_confirmacao_sem_sessao_nao_importa(admin_client_logado, modulo):
    antes = contagens()
    resposta = admin_client_logado.post(URL_CONFIRMAR, {"token": "qualquer"})
    assert resposta.status_code == 302
    assert resposta.url == URL_UPLOAD
    assert contagens() == antes


def test_confirmar_duas_vezes_nao_duplica(admin_client_logado, modulo, outro_modulo):
    enviar(admin_client_logado, csv(CSV_EXEMPLO))
    token = admin_client_logado.get(URL_PREVIEW).context["token"]

    admin_client_logado.post(URL_CONFIRMAR, {"token": token})
    depois_da_primeira = contagens()

    # A sessao foi limpa; a segunda tentativa nao encontra lote.
    admin_client_logado.post(URL_CONFIRMAR, {"token": token})
    assert contagens() == depois_da_primeira


def test_cancelar_descarta_o_lote(admin_client_logado, modulo, outro_modulo):
    enviar(admin_client_logado, csv(CSV_EXEMPLO))
    antes = contagens()

    resposta = admin_client_logado.post(URL_CANCELAR)
    assert resposta.status_code == 302
    assert contagens() == antes

    # Depois de cancelar, nao ha mais preview para confirmar.
    assert admin_client_logado.get(URL_PREVIEW).url == URL_UPLOAD


def test_importacao_por_xlsx_pelas_views(admin_client_logado, modulo):
    arquivo = planilha(
        [["nome", "email", "modulo"], ["Ana Souza", "ana.souza@exemplo.test", "MOD1"]]
    )
    enviar(admin_client_logado, arquivo)
    token = admin_client_logado.get(URL_PREVIEW).context["token"]
    admin_client_logado.post(URL_CONFIRMAR, {"token": token})

    aluno = User.objects.get(email="ana.souza@exemplo.test")
    assert aluno.student_profile.source == StudentSource.IMPORT
    assert aluno.enrollments.filter(module=modulo).exists()


def test_upload_sem_senha_padrao_configurada_nao_importa(
    admin_client_logado, modulo, settings
):
    enviar(admin_client_logado, csv("nome,email,modulo\nAna,ana@exemplo.test,MOD1\n"))
    token = admin_client_logado.get(URL_PREVIEW).context["token"]

    settings.DEFAULT_STUDENT_PASSWORD = ""
    antes = contagens()
    admin_client_logado.post(URL_CONFIRMAR, {"token": token})
    assert contagens() == antes


# ---------------------------------------------------------------------------
# Atomicidade
# ---------------------------------------------------------------------------


def test_falha_no_meio_do_lote_desfaz_tudo(admin_client_logado, modulo, outro_modulo):
    """
    O lote inteiro roda em uma transacao.

    Se a criacao de uma matricula falhar no meio do caminho, nada pode restar
    gravado: uma importacao pela metade seria pior que nenhuma, porque o
    administrador nao teria como saber onde ela parou.
    """
    enviar(admin_client_logado, csv(CSV_EXEMPLO))
    token = admin_client_logado.get(URL_PREVIEW).context["token"]
    antes = contagens()

    original = Enrollment.objects.create
    chamadas = {"n": 0}

    def falhar_na_segunda(*args, **kwargs):
        chamadas["n"] += 1
        if chamadas["n"] == 2:
            raise RuntimeError("falha simulada no meio do lote")
        return original(*args, **kwargs)

    with mock.patch.object(
        Enrollment.objects, "create", side_effect=falhar_na_segunda
    ):
        with pytest.raises(RuntimeError):
            admin_client_logado.post(URL_CONFIRMAR, {"token": token})

    assert contagens() == antes


# ---------------------------------------------------------------------------
# Segredos e auditoria
# ---------------------------------------------------------------------------


def test_senha_padrao_nao_aparece_nas_telas_de_importacao(
    admin_client_logado, modulo, outro_modulo, senha_padrao
):
    assert senha_padrao not in admin_client_logado.get(URL_UPLOAD).content.decode()

    enviar(admin_client_logado, csv(CSV_EXEMPLO))
    assert senha_padrao not in admin_client_logado.get(URL_PREVIEW).content.decode()


def test_auditoria_da_importacao(admin_client_logado, modulo, outro_modulo, senha_padrao):
    enviar(admin_client_logado, csv(CSV_EXEMPLO))
    token = admin_client_logado.get(URL_PREVIEW).context["token"]
    admin_client_logado.post(URL_CONFIRMAR, {"token": token})

    assert AuditLog.objects.filter(event=AuditEvent.STUDENT_CREATED).count() == 3
    assert AuditLog.objects.filter(event=AuditEvent.ENROLLMENT_CREATED).count() == 4

    lote = AuditLog.objects.get(event=AuditEvent.STUDENT_IMPORT_COMPLETED)
    assert lote.metadata == {
        "rows_received": 6,
        "students_created": 3,
        "enrollments_created": 4,
        "skipped": 1,
        "invalid": 1,
    }
    # Minimizacao de dados: o resumo do lote nao carrega a lista de e-mails.
    assert "joao@example.com" not in str(lote.metadata)


def test_senha_padrao_nao_chega_a_auditoria(
    admin_client_logado, modulo, outro_modulo, senha_padrao
):
    enviar(admin_client_logado, csv(CSV_EXEMPLO))
    token = admin_client_logado.get(URL_PREVIEW).context["token"]
    admin_client_logado.post(URL_CONFIRMAR, {"token": token})

    trilha = " ".join(str(log.metadata) for log in AuditLog.objects.all())
    assert senha_padrao not in trilha


# ---------------------------------------------------------------------------
# Autorizacao
# ---------------------------------------------------------------------------


def test_aluno_nao_acessa_a_importacao(student_client_logado, modulo):
    assert student_client_logado.get(URL_UPLOAD).status_code == 403
    assert student_client_logado.get(URL_PREVIEW).status_code == 403
    assert student_client_logado.post(URL_CONFIRMAR, {"token": "x"}).status_code == 403
    assert student_client_logado.post(URL_CANCELAR).status_code == 403


def test_anonimo_e_redirecionado_na_importacao(client):
    resposta = client.get(URL_UPLOAD)
    assert resposta.status_code == 302
    assert "/login/" in resposta.url


@pytest.mark.parametrize("url", [URL_CONFIRMAR, URL_CANCELAR])
def test_confirmacao_e_cancelamento_recusam_get(admin_client_logado, url):
    assert admin_client_logado.get(url).status_code == 405
